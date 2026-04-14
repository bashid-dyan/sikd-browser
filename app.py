"""
SIKD Browser — Flask Web App
Jelajahi data APBD dari DJPK Kemenkeu
"""

import os
import io
import atexit
from flask import Flask, request, jsonify, render_template, send_file
import sikd_client

app = Flask(__name__, static_folder='static', static_url_path='/static')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/provinsi')
def api_provinsi():
    tahun = request.args.get('tahun', '2025')
    try:
        data = sikd_client.get_provinsi(tahun)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pemda')
def api_pemda():
    provinsi = request.args.get('provinsi', '')
    tahun = request.args.get('tahun', '2025')
    if not provinsi:
        return jsonify({'error': 'Parameter provinsi harus diisi'}), 400
    try:
        data = sikd_client.get_pemda(provinsi, tahun)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/apbd')
def api_apbd():
    tahun = request.args.get('tahun', '2025')
    provinsi = request.args.get('provinsi', '--')
    pemda = request.args.get('pemda', '--')
    periode = request.args.get('periode', '12')
    try:
        data = sikd_client.get_apbd(tahun, provinsi, pemda, periode)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/compare')
def api_compare():
    tahun_str = request.args.get('tahun', '')  # comma-separated: 2023,2024,2025
    provinsi = request.args.get('provinsi', '--')
    pemda = request.args.get('pemda', '--')
    periode = request.args.get('periode', '12')

    if not tahun_str:
        return jsonify({'error': 'Parameter tahun harus diisi (misal: 2023,2024,2025)'}), 400

    tahun_list = [t.strip() for t in tahun_str.split(',') if t.strip()]
    try:
        data = sikd_client.get_apbd_compare(tahun_list, provinsi, pemda, periode)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download-excel')
def api_download_excel():
    """Download data APBD sebagai file Excel asli."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tahun = request.args.get('tahun', '2025')
    provinsi = request.args.get('provinsi', '--')
    pemda = request.args.get('pemda', '--')
    periode = request.args.get('periode', '12')
    mode = request.args.get('mode', 'single')  # 'single' atau 'compare'
    tahun_list_str = request.args.get('tahun_list', '')
    filename = request.args.get('filename', 'APBD')

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styles
    hdr_fill = PatternFill(start_color="4c6ef5", end_color="4c6ef5", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    lvl0_fill = PatternFill(start_color="DBE4FF", end_color="DBE4FF", fill_type="solid")
    lvl0_font = Font(bold=True, size=10, name="Arial")
    lvl1_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    lvl1_font = Font(bold=True, size=10, name="Arial")
    data_font = Font(size=10, name="Arial")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if mode == 'compare' and tahun_list_str:
        # Compare mode
        tahun_list = [t.strip() for t in tahun_list_str.split(',') if t.strip()]
        ws.title = "Perbandingan APBD"

        all_data = sikd_client.get_apbd_compare(tahun_list, provinsi, pemda, periode)
        years = sorted(all_data.keys())

        # Collect all unique akun
        all_akun = []
        akun_seen = set()
        for y in years:
            for r in all_data[y].get('rows', []):
                if r['akun'] not in akun_seen:
                    akun_seen.add(r['akun'])
                    all_akun.append({'akun': r['akun'], 'level': r['level']})

        # Header
        headers = ['Akun']
        for y in years:
            headers += [f'Anggaran {y}', f'Realisasi {y}', f'% {y}']
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.fill, c.font, c.border = hdr_fill, hdr_font, thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Rows
        for i, a in enumerate(all_akun, 2):
            ws.cell(row=i, column=1, value=a['akun']).border = thin_border
            col = 2
            for y in years:
                row_data = next((r for r in all_data[y].get('rows', []) if r['akun'] == a['akun']), None)
                if row_data:
                    ws.cell(row=i, column=col, value=row_data.get('anggaran', 0)).border = thin_border
                    ws.cell(row=i, column=col+1, value=row_data.get('realisasi', 0)).border = thin_border
                    try:
                        pct = float(row_data.get('persentase', '0'))
                        ws.cell(row=i, column=col+2, value=pct/100).border = thin_border
                        ws.cell(row=i, column=col+2).number_format = '0.00%'
                    except (ValueError, TypeError):
                        ws.cell(row=i, column=col+2, value=row_data.get('persentase', '')).border = thin_border
                else:
                    for offset in range(3):
                        ws.cell(row=i, column=col+offset, value='-').border = thin_border
                col += 3

            # Apply styling based on level
            for col_idx in range(1, len(headers)+1):
                cell = ws.cell(row=i, column=col_idx)
                cell.font = data_font
                if a['level'] == 0:
                    cell.fill = lvl0_fill
                    cell.font = lvl0_font
                elif a['level'] == 1:
                    cell.fill = lvl1_fill
                    cell.font = lvl1_font
                if col_idx >= 2 and col_idx % 3 != 1:  # Number columns
                    cell.number_format = '#,##0'

        # Column widths
        ws.column_dimensions['A'].width = 50
        for i in range(len(years)):
            col_letter_base = openpyxl.utils.get_column_letter(2 + i*3)
            ws.column_dimensions[col_letter_base].width = 18
            ws.column_dimensions[openpyxl.utils.get_column_letter(3 + i*3)].width = 18
            ws.column_dimensions[openpyxl.utils.get_column_letter(4 + i*3)].width = 10

    else:
        # Single year mode
        ws.title = f"APBD {tahun}"

        data = sikd_client.get_apbd(tahun, provinsi, pemda, periode)
        rows = data.get('rows', [])

        headers = ['Akun', 'Anggaran', 'Realisasi', '%']
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.fill, c.font, c.border = hdr_fill, hdr_font, thin_border
            c.alignment = Alignment(horizontal='center', vertical='center')

        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r.get('akun', '')).border = thin_border
            ws.cell(row=i, column=2, value=r.get('anggaran', 0)).border = thin_border
            ws.cell(row=i, column=3, value=r.get('realisasi', 0)).border = thin_border
            try:
                pct = float(r.get('persentase', '0'))
                ws.cell(row=i, column=4, value=pct/100).border = thin_border
                ws.cell(row=i, column=4).number_format = '0.00%'
            except (ValueError, TypeError):
                ws.cell(row=i, column=4, value=r.get('persentase', '')).border = thin_border

            # Styling by level
            for col_idx in range(1, 5):
                cell = ws.cell(row=i, column=col_idx)
                cell.font = data_font
                if r.get('level') == 0:
                    cell.fill = lvl0_fill
                    cell.font = lvl0_font
                elif r.get('level') == 1:
                    cell.fill = lvl1_fill
                    cell.font = lvl1_font
                if col_idx in (2, 3):
                    cell.number_format = '#,##0'

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{filename}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


atexit.register(sikd_client.close)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
